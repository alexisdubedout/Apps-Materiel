#!/usr/bin/env python3
import sys
from openpyxl import load_workbook
from openpyxl.pivot.table import TableDefinition, PivotTable
from openpyxl.pivot.cache import CacheDefinition
from openpyxl.pivot.fields import RowFields, ColumnFields, DataField, PageField

def add_pivot_table(filepath):
    wb = load_workbook(filepath)
    
    # Trouver la feuille qui commence par "Export"
    ws = None
    for sheet in wb.worksheets:
        if sheet.title.startswith('Export'):
            ws = sheet
            break
    
    if not ws:
        print("❌ Feuille Export introuvable")
        return
    
    # Déterminer la plage de données
    max_row = ws.max_row
    data_range = f'A1:I{max_row}'
    
    # Créer le cache de données
    cache = CacheDefinition()
    cache.cacheSource.worksheetSource.ref = data_range
    cache.cacheSource.worksheetSource.sheet = ws.title
    
    # Créer le TCD
    pivot = PivotTable()
    pivot.location.ref = f'K1'
    pivot.location.firstHeaderRow = 1
    pivot.location.firstDataRow = 1
    pivot.location.firstDataCol = 1
    
    # Filtre de page : Code article (colonne 3)
    pivot.pageFields.append(PageField(fld=2, name='Code article'))
    
    # Lignes : Type emplacement (8), Détail (9), Description emplacement (7)
    pivot.rowFields = RowFields()
    pivot.rowFields.field.append({'x': 7})  # Type emplacement
    pivot.rowFields.field.append({'x': 8})  # Détail
    pivot.rowFields.field.append({'x': 6})  # Description emplacement
    
    # Colonnes : Code produit (4)
    pivot.colFields = ColumnFields()
    pivot.colFields.field.append({'x': 3})  # Code produit
    
    # Valeurs : Compte de Code actif
    pivot.dataFields.append(DataField(fld=0, name='Nombre de Code actif', baseField=0, baseItem=0))
    
    ws.add_pivot(pivot, data_range)
    
    wb.save(filepath)
    print(f"✅ TCD ajouté à {filepath}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python add-pivot-table.py <filepath>")
        sys.exit(1)
    
    add_pivot_table(sys.argv[1])
